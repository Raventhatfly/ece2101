from openpyxl import Workbook, load_workbook
import os
import numpy as np

path1 = "2101\活动记实"
class_dict = {}
for root, dirs, files in os.walk(path1):
    for file in files:
        if file.endswith("xlsx"):
            name = root.split('\\')[-1]
            wb = load_workbook(os.path.join(root,file))
            sheet = wb.active
            column_data = []
            for cell in sheet['H']:
                value = sheet["H" + str(cell.row)].value
                if(sheet[str(cell.row)][0].value != "示例" and type(value) == int):
                    column_data.append(value)
            column_data = np.array(column_data)
            class_dict[name] = column_data.sum()
            wb.close()

path2 = "附件3：ZJUI思想政治素质评价班级汇总表（2021-2022学年）.xlsx"
wb = load_workbook(path2)
sheet = wb.active
row_map = {}
for cell in sheet["C"]:
    if type(cell.value)==str:
        row_map[cell.value] = cell.row 
wb.close()

wb = load_workbook(path2)
sheet = wb.active
for name in class_dict:
    row = row_map[name]
    sheet["D" + str(row)] = class_dict[name]
wb.save(path2)

print(class_dict)
